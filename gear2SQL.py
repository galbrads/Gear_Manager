from PySide import QtGui, QtCore
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string
import database as datab
import Util
import sys
import datetime

app = QtGui.QApplication(sys.argv)

db = datab.Database(None, 'UCMC_DB.sqlite')

wb = load_workbook(filename='UCMC Inventory.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name(name='Inventory')

# Read and confirm the header
fHead = []
for row in ws.iter_rows():
    gearAttr = {}
    for cell in row:
        if not cell.column:
            continue

        col = column_index_from_string(cell.column)
        if cell.row > 1 and cell.value:
            # Read the remaining members
            gearAttr[fHead[col - 1]] = cell.value

        elif cell.row == 1 and cell.value:
            fHead.append(cell.value)

    if gearAttr and 'Name' in gearAttr.keys() and gearAttr['Name']:

        gearAttr['Name'] = gearAttr['Name'].encode("ascii", "ignore")

        if 'Bar codeID' in gearAttr.keys():
            if isinstance(gearAttr['Bar codeID'], (unicode, basestring)):
                gearAttr['Bar codeID'] = gearAttr['Bar codeID'].strip()

            if gearAttr['Bar codeID']:
                gearAttr['ID'] = str(gearAttr['Bar codeID']).strip().split('.')[0]

        if 'Item #' in gearAttr.keys():
            if isinstance(gearAttr['Item #'], (unicode, basestring)):
                gearAttr['Item #'] = gearAttr['Item #'].strip()

            if gearAttr['Item #']:
                gearAttr['ID'] = str(gearAttr['Item #'].strip())

        if 'PurchasePrice' in gearAttr.keys():
            if isinstance(gearAttr['PurchasePrice'], (basestring, unicode)):
                gearAttr['PurchasePrice'] = gearAttr['PurchasePrice'].strip().replace('..', '.')

            if gearAttr['PurchasePrice']:
                gearAttr['Price'] = gearAttr['PurchasePrice']
            else:
                gearAttr['Price'] = 0.0
        else:
            gearAttr['Price'] = 0.0

        if 'Purchase Date' in gearAttr.keys() and isinstance(gearAttr['Purchase Date'], datetime.datetime):

            if gearAttr['Purchase Date']:
                gearAttr['PurchaseDate'] = QtCore.QDateTime(gearAttr['Purchase Date'])
            else:
                gearAttr['PurchaseDate'] = QtCore.QDate(1900, 1, 1)

            gearAttr['PurchaseDate'] = Util.convert_date('Qt2DB', gearAttr['PurchaseDate'])

        if 'Manufacturer' in gearAttr.keys() and not gearAttr['Manufacturer']:
            gearAttr['Manufacturer'] = ''

        if 'Category' in gearAttr.keys():
            if gearAttr['Category']:
                gearAttr['Category'] = int(gearAttr['Category'])
            else:
                gearAttr['Category'] = -1

        if 'Miscellaneous' in gearAttr.keys() and gearAttr['Miscellaneous']:
            gearAttr['Misc'] = gearAttr['Miscellaneous'].encode("ascii", "ignore")

        gear = db.getGear(gearAttr['ID'])
        if gear:
            db.updateItem('Gear', gearAttr, gear)
        else:
            db.addItem('Gear', gearAttr)
