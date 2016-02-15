from PySide import QtGui, QtCore
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string
import database as datab
import datetime
import Util
import sys
import os

# Semester Start/End dates
fallS = Util.convert_date('Qt2DB', QtCore.QDate(2014, 8, 25))
fallE = Util.convert_date('Qt2DB', QtCore.QDate(2014, 12, 13))
sprS = Util .convert_date('Qt2DB', QtCore.QDate(2015, 1, 12))
sprE = Util .convert_date('Qt2DB', QtCore.QDate(2015, 5, 2))
sumS = Util .convert_date('Qt2DB', QtCore.QDate(2015, 5, 11))
sumE = Util .convert_date('Qt2DB', QtCore.QDate(2015, 8, 8))

rentalFee = 20

today = Util.convert_date('Qt2DB', QtCore.QDate.currentDate())

app = QtGui.QApplication(sys.argv)

dbName = 'UCMC_DB.sqlite'

if os.path.isfile(dbName):
    print 'Removing', dbName
    os.remove(dbName)

db = datab.Database(None, dbName)

# Initialize the admin table
db.execQuery("DELETE FROM Admin", 'member2SQL.py')
db.execQuery("VACUUM", 'member2SQL.py')
db.execQuery("INSERT INTO Admin VALUES({}, {}, {:0.2f}, '{}', '{}', '{}', '{}', '{}', '{}')".format(3, 1, rentalFee, fallS, fallE, sprS, sprE, sumS, sumE), 'member2SQL.py')

wb = load_workbook(filename='UCMC Member List.xlsx', use_iterators=True)
for ws in wb:

    print ws

    # Read and confirm the header
    fHead = []
    for row in ws.iter_rows():
        membAttr = {}
        for cell in row:
            if not cell.column:
                continue

            col = column_index_from_string(cell.column)
            if cell.row > 1 and col <= len(fHead) + 1:
                # Read the remaining members
                membAttr[fHead[col - 1]] = cell.value
#                 if isinstance(cell.value, (basestring, unicode)):
#                     cell.value = cell.value.strip()

            elif cell.row == 1 and cell.value:
                fHead.append(cell.value.strip())

        if membAttr and 'First Name' in membAttr.keys() and membAttr['First Name']:

            membAttr['FirstName'] = membAttr['First Name'].encode("ascii", "ignore").strip()
            membAttr['LastName'] = membAttr['Last Name'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['Phone'], (basestring, unicode)):

                membAttr['Phone'] = membAttr['Phone'].encode("ascii", "ignore").strip()

            membAttr['StudentID'] = ''
            if isinstance(membAttr['UC #'], (basestring, unicode)):

                MID = membAttr['UC #'].encode("ascii", "ignore").strip()

                if len(MID) == 9:
                    membAttr['StudentID'] = MID
                else:
                    membAttr['Note'] = 'ID: ' + MID

            if isinstance(membAttr['Email'], (basestring, unicode)):
                membAttr['Email'] = membAttr['Email'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['Phone'], (basestring, unicode)):
                membAttr['Phone'] = membAttr['Phone'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['Address'], (basestring, unicode)):

                membAttr['Address'] = membAttr['Address'].encode("ascii", "ignore").strip()

                if membAttr['Address'] == 'Smurf House':
                    membAttr['Street'] = membAttr['Address']
                else:
                    membAttr['Address'] = membAttr['Address'].replace(',', ', ').split()
                    print membAttr['LastName'], membAttr['FirstName'], membAttr['Address']
                    membAttr['Street'] = ' '.join(membAttr['Address'][:-3]).strip(',').strip().encode("ascii", "ignore")
                    membAttr['City'  ] = membAttr['Address'][-3].strip(',').strip().encode("ascii", "ignore")
                    membAttr['State' ] = membAttr['Address'][-2].strip(',').strip().encode("ascii", "ignore").upper()
                    membAttr['Zip'   ] = membAttr['Address'][-1].strip(',').strip().encode("ascii", "ignore")

            if isinstance(membAttr['Em Contact'], (basestring, unicode)):

                membAttr['EmName'] = membAttr['Em Contact'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['Em Phone'], (basestring, unicode)):

                membAttr['Em Phone'] = membAttr['Em Phone'].upper().replace('H', ' H').replace('W', ' W').replace('C', ' C').strip()

                for phone in membAttr['Em Phone'].split():
                    pType = phone[0].upper().strip()
                    if pType.isalpha():
                        phone = phone[1:].strip().encode("ascii", "ignore")
                        membAttr['EmPhone' + pType] = phone.strip()
                    else:
                        membAttr['EmPhoneH'] = phone.strip()

            if isinstance(membAttr['Rel'], (basestring, unicode)):

                membAttr['EmRel'] = membAttr['Rel'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['B-date'], datetime.datetime):
                membAttr['B-date'] = str(membAttr['B-date']).split()[0]

            if isinstance(membAttr['B-date'], (basestring, unicode)):

                membAttr['Birthday'] = QtCore.QDate.fromString(membAttr['B-date'], 'yyyy-MM-dd').toString(Util.DB_Date)

            if isinstance(membAttr['Med Cond'], (basestring, unicode)):

                membAttr['Med'] = membAttr['Med Cond'].encode("ascii", "ignore").strip()

            for key in membAttr.keys():
                if not membAttr[key]:
                    membAttr[key] = ''

            if 'Birthday' not in membAttr.keys() or not membAttr['Birthday']:

                query = db.getQuery('Birthday', 'Member', searchAtt=['FirstName', 'LastName'],
                                    search=[membAttr['FirstName'], membAttr['LastName']])

                query.first()
                membAttr['Birthday'] = query.value(0)

                if not membAttr['Birthday']:
                    membAttr['Birthday'] = '1900-01-01'

            # If the forms were from this year, mark them as waviers on file
            if ws.title == '2014-2015':
                membAttr['FormsCurrent'] = 1
                membAttr['FormsDate'] = today

            memb = db.getMember(membAttr['FirstName'], membAttr['LastName'], membAttr['Birthday'])
            if memb:
                db.updateItem('Member', membAttr, memb)
            else:
                db.addItem('Member', membAttr)

wb = load_workbook(filename='UCMC Member List 2.xlsx', use_iterators=True)

for sheetName in ['2015 Yr']:

    ws = wb.get_sheet_by_name(name=sheetName)

    # Read and confirm the header
    fHead = []
    for row in ws.iter_rows():
        membAttr = {}
        for cell in row:
            if not cell.column or cell.column == 'A':
                continue

            col = column_index_from_string(cell.column)
            if cell.row > 1 and col <= len(fHead) + 1:
                # Read the remaining members
                membAttr[fHead[col - 2]] = cell.value

            elif cell.row == 1 and cell.value:
                fHead.append(cell.value)

        if membAttr and 'First Name' in membAttr.keys() and membAttr['First Name']:

            membAttr['FirstName'] = membAttr['First Name'].encode("ascii", "ignore").strip()
            membAttr['LastName'] = membAttr['Last Name'].encode("ascii", "ignore").strip()

            if isinstance(membAttr['Wvr'], (basestring, unicode)):
                membAttr['Wvr'] = membAttr['Wvr'].strip()
            if membAttr['Wvr']:
                membAttr['FormsCurrent'] = 1
                membAttr['FormsDate'] = today
            else:
                membAttr['FormsCurrent'] = 0

            if isinstance(membAttr['CLnk'], (basestring, unicode)):
                membAttr['CLnk'] = membAttr['CLnk'].strip()
            if membAttr['CLnk']:
                membAttr['CampusLink'] = 1
                membAttr['CampusLinkDate'] = today
            else:
                membAttr['CampusLink'] = 0

            if isinstance(membAttr['Gear Policy'], (basestring, unicode)):
                membAttr['Gear Policy'] = membAttr['Gear Policy'].strip()
            if membAttr['Gear Policy']:
                membAttr['Note'] = membAttr['Gear Policy'].encode("ascii", "ignore").strip()
                if 'L' in membAttr['Note'].upper() and 'CLIMB' in membAttr['Note'].upper():
                    membAttr[Util.certifications[0] + 'Cert'] = 1
                    if 'TRAD' in membAttr['Note'].upper():
                        membAttr[Util.certifications[1] + 'Cert'] = 1
                    if 'ICE' in membAttr['Note'].upper():
                        membAttr[Util.certifications[2] + 'Cert'] = 1

            if isinstance(membAttr['Phone'], (basestring, unicode)):
                membAttr['Phone'] = membAttr['Phone'].strip()
            if membAttr['Phone']:
                membAttr['Phone'] = membAttr['Phone'].encode("ascii", "ignore").strip()
            else:
                membAttr['Phone'] = ''

            for key in membAttr.keys():
                if membAttr[key] is None or not membAttr[key] and not Util.is_number(membAttr[key]):
                    membAttr[key] = ''

            membAttr['1st Sem'] = membAttr['1st Sem'].strip().upper()
            membAttr['2nd Sem'] = membAttr['2nd Sem'].strip().upper()
            membAttr['Sum'    ] = membAttr['Sum'    ].strip().upper()

            if membAttr['1st Sem'] or membAttr['2nd Sem'] or membAttr['Sum']:
                membAttr['MembStat'] = ' '.join([membAttr['1st Sem'], membAttr['2nd Sem'], membAttr['Sum']])
                if 'LIFE' in membAttr['MembStat']:
                    membAttr['MembStat'] = db.memberStatus[-1]
                elif 'EXC' in membAttr['MembStat']:
                    membAttr['MembStat'] = db.memberStatus[2]
                else:
                    membAttr['MembStat'] = db.memberStatus[0]
            else:
                membAttr['MembStat'] = db.memberStatus[0]

            query = db.getQuery('Birthday', 'Member', searchAtt=['FirstName', 'LastName'],
                                search=[membAttr['FirstName'], membAttr['LastName']])

            query.first()
            membAttr['Birthday'] = query.value(0)

            if membAttr['Birthday']:

                memb = db.getMember(membAttr['FirstName'], membAttr['LastName'], membAttr['Birthday'])

                if memb:
                    for key in membAttr.keys():
                        if not membAttr[key] and key in memb.__dict__.keys() and memb.__dict__[key]:
                            membAttr[key] = memb.__dict__[key]

                    db.updateItem('Member', membAttr, memb)

                    attrList = {'MID': memb.ID, 'Type': 'Rental', 'Amount': '{:0.2f}'.format(rentalFee), 'Comment': ''}

                    if 'PD' in membAttr['1st Sem']:
                        attrList['Date'] = fallS
                        db.addItem('FinancialTrans', attrList)
                    if 'PD' in membAttr['2nd Sem']:
                        attrList['Date'] = sprS
                        db.addItem('FinancialTrans', attrList)
                    if 'PD' in membAttr['Sum']:
                        attrList['Date'] = sumS
                        db.addItem('FinancialTrans', attrList)

wb = load_workbook(filename='UCMC Inventory.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name(name='Gear')

# Read and confirm the header
fHead = []
for row in ws.iter_rows():
    gearAttr = {}
    for cell in row:
        if not cell.column:
            continue

        col = column_index_from_string(cell.column)
        if cell.row > 1 and cell.value:
            # Read the remaining members
            gearAttr[fHead[col - 1]] = cell.value

        elif cell.row == 1 and cell.value:
            fHead.append(cell.value)

    if gearAttr and 'Name' in gearAttr.keys() and gearAttr['Name']:

        gearAttr['Name'] = gearAttr['Name'].encode("ascii", "ignore").strip()

        if 'Bar codeID' in gearAttr.keys():
            if isinstance(gearAttr['Bar codeID'], (unicode, basestring)):
                gearAttr['Bar codeID'] = gearAttr['Bar codeID'].strip()

            if gearAttr['Bar codeID']:
                gearAttr['ID'] = str(gearAttr['Bar codeID']).strip().split('.')[0]

        if 'Item #' in gearAttr.keys():
            if isinstance(gearAttr['Item #'], (unicode, basestring)):
                gearAttr['Item #'] = gearAttr['Item #'].strip()

            if gearAttr['Item #']:
                gearAttr['ID'] = str(gearAttr['Item #'].strip())

        gearAttr['Price'] = 0.0
        if 'PurchasePrice' in gearAttr.keys():
            if isinstance(gearAttr['PurchasePrice'], (basestring, unicode)):
                gearAttr['PurchasePrice'] = gearAttr['PurchasePrice'].strip().replace('..', '.')

            if gearAttr['PurchasePrice']:
                gearAttr['Price'] = gearAttr['PurchasePrice']

        if 'Purchase Date' in gearAttr.keys() and isinstance(gearAttr['Purchase Date'], datetime.datetime):

            if gearAttr['Purchase Date']:
                gearAttr['PurchaseDate'] = QtCore.QDateTime(gearAttr['Purchase Date'])
            else:
                gearAttr['PurchaseDate'] = QtCore.QDate(1900, 1, 1)

            gearAttr['PurchaseDate'] = Util.convert_date('Qt2DB', gearAttr['PurchaseDate'])

        if 'Manufacturer' in gearAttr.keys() and not gearAttr['Manufacturer']:
            gearAttr['Manufacturer'] = ''

        if 'Category' in gearAttr.keys():
            if gearAttr['Category']:
                gearAttr['Category'] = int(gearAttr['Category'])
            else:
                gearAttr['Category'] = -1

        if 'Miscellaneous' in gearAttr.keys() and gearAttr['Miscellaneous']:
            gearAttr['Misc'] = gearAttr['Miscellaneous'].encode("ascii", "ignore").strip()

        gear = db.getGear(gearAttr['ID'])
        # # print gearAttr
        if gear:
            db.updateItem('Gear', gearAttr, gear)
        else:
            db.addItem('Gear', gearAttr)

db.close()
