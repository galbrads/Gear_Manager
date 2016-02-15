from PySide import QtSql, QtCore, QtGui
from openpyxl import Workbook, load_workbook
import datetime
import Util
import shutil
import os

tableMember = None
tableGear = None
tableTransaction = None
tableArchiveTrans = None
tableAdmin = None
tableFinancialTran = None
tableGearMaintenance = None

fullFirstLast = None
gearNameIDList = None


class QueryError(Exception):

    def __init__(self, query, loc):
        self.query = query
        self.loc = loc

    def __str__(self):

        errorMessage = '''\n        
Error   : {e}
          {t}
Location: {l}
Query   : {q}
'''.format(e=self.query.lastError(),
           t=self.query.lastError().type(),
           q=self.query.lastQuery(),
           l=self.loc)

        return errorMessage


class DictList(list):

    def __init__(self, item_list):

        self._list = []
        self._dict = {}

        for n in xrange(len(item_list)):
            self._dict[item_list[n]] = n
            self._list.append(item_list[n])

    def __getitem__(self, key):

        if isinstance(key, int):
            return self._list[key]
        elif isinstance(key, (basestring, unicode)):
            return self._dict[key]
        else:
            raise ValueError

    def getList(self):

        return self._list

    def getDict(self):

        return self._dict

    def nCerts(self):

        return len(self._list)


class QueryTools(object):

    def execQuery(self, cmd, functionNameLoc):

        query = QtSql.QSqlQuery(cmd)

        if query.lastError().type() != QtSql.QSqlError.NoError:
            raise QueryError(query, functionNameLoc)

        return query

    def getQuery(self, findAttr, table, searchAtt=None, search=None, sort=None):

        cmd = 'SELECT {0} FROM {1}'.format(findAttr, table)

        if isinstance(searchAtt, list) and isinstance(search, list):

            if searchAtt and search:
                for n in xrange(len(searchAtt)):
                    if n == 0: cmd += ' WHERE'
                    if n > 0: cmd += ' AND'

                    if isinstance(search[n], (basestring, unicode)):
                        search[n] = '"{0}"'.format(search[n].replace('"', '""'))
                    cmd += ' {0}={1}'.format(searchAtt[n], search[n])

        elif isinstance(searchAtt, (basestring, unicode)) and isinstance(search, (basestring, unicode)):
            cmd += ' WHERE {0}="{1}"'.format(searchAtt, search.replace('"', '""'))

        if sort:
            cmd += ' {0}'.format(sort)

        query = self.execQuery(cmd, 'database.py -> getQuery')

        return query

    def setFieldTo(self, table, attr, val):

        # If val is a string, append " to beginning and end
        if isinstance(val, (basestring, unicode)):
            val = '"{0}"'.format(val)

        cmd = 'UPDATE {0} SET {1}={2}'.format(table, attr, val)
        self.execQuery(cmd, 'database.py -> querySetFieldTo')

    def qLen(self, query):

        # Copy the query
        q = query

        # Set the query at the end of the query
        q.last()

        # .at() + 1, when on the last index of the query is the length of the query
        lenQ = q.at() + 1

        return lenQ

    def getSemesterDates(self):

        semNames = ['SemFallStart', 'SemFallEnd', 'SemSprStart', 'SemSprEnd', 'SemSumStart', 'SemSumEnd']
        semDates = {}

        query = self.getQuery(', '.join(semNames), 'Admin')

        query.first()
        for n, sName in enumerate(semNames):
            semDates[sName] = Util.convert_date('DB2Qt', query.value(n))

        return semDates


class Table(QueryTools):

    def __init__(self, entries):

        self.name = entries[0]
        self.fields = []

        cmd = 'CREATE TABLE IF NOT EXISTS {0} ('.format(self.name)

        for n in xrange(1, len(entries)):
            if isinstance(entries[n], list):
                self.fields.append(entries[n][0])
                cmd += '{0} {1}, '.format(entries[n][0], entries[n][1])
        cmd += ')'
        if not isinstance(entries[-1], list):
            cmd = cmd.replace(', )', ' {0})'.format(entries[-1]))
        else:
            cmd = cmd.replace(', )', ')')

        self.execQuery(cmd, 'database.py -> Table -> __init__')


class TableBase(Table):

    def __init__(self):

        # self.CertDef = DictList(['LSport', 'LTrad', 'LIce', 'KayakRoll'])
        self.paymentType = DictList(['Rental', 'Late Fee', 'Damages', 'Refund', 'Other'])
        self.memberStatus = DictList(['Regular', 'President', 'Vice President',
                                      'Treasurer', 'Secretary', 'Gear Manager', 'Honorary'])
        self.studentStatus = DictList(['Undergrad', 'Graduate', 'Alumni', 'Non-Student'])
        self.gear_category = DictList([str(n) for n in range(1, 9)])

    def table_entry(self, table, *args):

        if table and args:

            attr = []
            search = []
            for n in xrange(0, len(args) / 2):
                attr.append(args[n])
                search.append(args[len(args) / 2 + n])

            # If an empty field is found, there will be no member match. return None
            if isinstance(search, (list, tuple)):
                for entry in search:
                    if not entry:
                        return None

            # Replace " with ""
            for n in xrange(len(search)):
                if isinstance(search[n], (basestring, unicode)):
                    search[n] = search[n].replace('"', '""')

            cmd = 'SELECT * FROM {0} WHERE {1}="{2}"'.format(table.name, attr[0], search[0])

            for i in xrange(1, len(attr)):
                cmd += ' AND {0}="{1}"'.format(attr[i], search[i])

            query = self.execQuery(cmd, 'database.py -> TableBase -> tableEntry')

            if self.qLen(query) == 1:

                query.first()
                for field in table.fields:
                    fInd = query.record().indexOf(field)
                    self.__dict__[field] = query.value(fInd)

                    if 'DATETIME' in field.upper():
                        self.__dict__[field] = QtCore.QDateTime.fromString(self.__dict__[field], Util.DB_DateTime)
                    elif 'DATE' in field.upper() or 'BIRTHDAY' in field.upper():
                        self.__dict__[field] = Util.convert_date('DB2Qt', self.__dict__[field])

                return True

        return False

    @staticmethod
    def split_name(name):

        first_name, last_name = None, None

        if ',' in name:
            name = "{0} {1}".format(name.split(',')[1].strip(), name.split(',')[0].strip())

        if name and fullFirstLast:

            try:
                ind = [x.upper() for x in fullFirstLast[0]].index(name.upper())
                first_name = fullFirstLast[1][ind]
                last_name = fullFirstLast[2][ind]
            except ValueError:
                pass

        return first_name, last_name


class Member(TableBase):

    def __init__(self, *args):
        super(Member, self).__init__()

        self._isDefined = False

        # If args has empty entries, no search was provided, return None
        if not args:
            return

        if len(args) == 1:

            self._isDefined = self.table_entry(tableMember, 'ID', args[0])

        elif len(args) == 2 or len(args) == 3:

            if len(args) == 2:

                if isinstance(args[0], QtGui.QLineEdit):
                    names = args[0].text()
                elif isinstance(args[0], QtGui.QComboBox):
                    names = args[0].currentText()

                if isinstance(args[1], QtGui.QLineEdit):
                    birth_day = args[1].dateDB()
                elif isinstance(args[1], QtGui.QComboBox):
                    birth_day = Util.convert_date('Disp2DB', args[1].currentText())

                first_name, last_name = self.split_name(names)
            elif len(args) == 3:
                first_name, last_name = args[0], args[1]
                birth_day = args[2]

            if first_name and last_name and birth_day:

                self._isDefined = self.table_entry(tableMember,
                                                  'FirstName', 'LastName', 'Birthday',
                                                  first_name, last_name, birth_day)

    def __nonzero__(self):

        return self._isDefined

    def eligibleToCheckOut(self):

        return self.formsCurrent() and (self.campusLink() or self.campus_link_waived())

    def formsCurrent(self):

        return self.FormsDate >= self.getSemesterDates()['SemFallStart']

    def campusLink(self):

        return self.CampusLinkDate <= QtCore.QDate.currentDate()

    def campus_link_waived(self):

        return self.StudStat == self.studentStatus[2] or self.StudStat == self.studentStatus[3]

    def has_active_transactions(self):

        cmd = 'SELECT GID FROM Transactions WHERE MID={0}'.format(self.ID)
        query = self.execQuery(cmd, 'database.py -> Member -> hasActiveTransactions')

        return self.qLen(query) >= 1

    def hasReqCerts(self, gear):

        # Loop through all defined certifications
        for cert in Util.certifications:
            if self.__dict__[cert + 'Cert'] < gear.__dict__[cert + 'Cert']:
                return False

        return True

    def is_currently_paid(self):

        query = self.getQuery('Date', 'FinancialTrans', searchAtt=['MID', 'Type'], search=[self.ID, 'Rental'], sort='ORDER BY Date DESC')

        if query and self.qLen(query) >= 1:

            query.first()
            paid_date = Util.convert_date('DB2Qt', query.value(0))

            semester_dates = self.getSemesterDates()

            # Find out which semester is the current semester
            today = QtCore.QDate.currentDate()

            # Include the break following the current semester
            if semester_dates['SemFallStart'] <= today < semester_dates['SemSprStart']:
                if semester_dates['SemFallStart'] <= paid_date < semester_dates['SemSprStart']:
                    return True

            if semester_dates['SemSprStart'] <= today < semester_dates['SemSumStart']:
                if semester_dates['SemSprStart'] <= paid_date < semester_dates['SemSumStart']:
                    return True

            if semester_dates['SemSumStart'] <= today < semester_dates['SemSumEnd']:
                if semester_dates['SemSumStart'] <= paid_date < semester_dates['SemSumEnd']:
                    return True

        return False

    def full_name(self):

        return '{0} {1}'.format(self.FirstName, self.LastName)

    def nameBDay(self):

        return '{0}, {1}'.format(self.full_name(), Util.convert_date('Qt2Disp', self.Birthday))


class Gear(TableBase):

    def __init__(self, *args):
        super(Gear, self).__init__()

        self._isDefined = False

        # Get the ID from *args
        for arg in args:

            if isinstance(arg, QtGui.QComboBox):
                arg = arg.currentText()
            elif isinstance(arg, QtGui.QLineEdit):
                arg = arg.text()

            try:
                gearID = [s for s in gearNameIDList[1] if arg.upper() == s.upper()]
            except:
                gearID = args

            if gearID:
                global tableGear
                self._isDefined = self.table_entry(tableGear, 'ID', arg)

    def __nonzero__(self):

        return self._isDefined

    def is_checkoutable(self):

        return not self.Unrentable and self.numAvailable() >= 1

    def numInInventory(self):

        query = self.getQuery('Quantity', 'Gear', searchAtt='ID', search=self.ID)
        query.first()
        numInv = query.record().value(0)

        return numInv

    def numCheckedOut(self):

        MID_List = self.whoHasMe()

        return len(MID_List)

    def numAvailable(self):

        return self.numInInventory() - self.numCheckedOut()

    def whoHasMe(self):

        query = self.getQuery('MID', 'Transactions', searchAtt='GID', search=self.ID)

        MID_List = []
        while query.next():
            MID_List.append(query.value(0))

        return MID_List


class Trans(TableBase, list):

    def __init__(self, member=None, gear=None, transaction_id=None):
        super(Trans, self).__init__()

        if not member and not gear and transaction_id:

            self.appendTrans(transaction_id)

        elif (member or gear) and not transaction_id:

            cmd = 'SELECT TID FROM {0} WHERE '.format(tableTransaction.name)

            if member:
                cmd += ' {}={}'.format('MID', member.ID)
            if gear:
                if member:
                    cmd += ' AND'
                cmd += ' {}="{}"'.format('GID', gear.ID)
            cmd += ' ORDER BY DueDate DESC'

            query = self.execQuery(cmd, 'database.py -> getTrans')

            while query.next():
                transaction_id = query.value(0)
                self.appendTrans(transaction_id)

    def __nonzero__(self):

        return len(self) >= 1

    def appendTrans(self, TID):

        trans = TableBase()
        trans.table_entry(tableTransaction, 'TID', TID)

        self.append(trans)

    def hasMID(self, member_id):

        for trans in self:

            if trans.MID == member_id:
                return True

        return False

    def hasGID(self, gear_id):

        for trans in self:

            if trans.GID == gear_id:
                return True

        return False


class Database(TableBase):

    def __init__(self, parent, name_db=None):
        super(Database, self).__init__()

        self.parent = parent
        self.getMember = Member
        self.getGear = Gear
        self.getTrans = Trans

        if name_db:

            self.nameDB = name_db

            self.SQLDB = QtSql.QSqlDatabase.addDatabase('QSQLITE')
            self.SQLDB.setDatabaseName(self.nameDB)
            self.SQLDB.open()

            # Check to see that the database was opened
            if not self.SQLDB.isOpen():
                print 'ERROR opening the database!'
                raise QtSql.QSqlError.ConnectionError

            def certDef(certification):

                return [certification + 'Cert', 'INTEGER DEFAULT 0']

            def mkCertFields(certification):

                return [certDef(certification),
                        [certification + 'CertDate', 'TEXT'],
                        [certification + 'CertVouch', 'TEXT']]

            self.tableMemberDef = ['Member',
                ['LastName'      , 'TEXT'],
                ['FirstName'     , 'TEXT'],
                ['ID'            , 'INTEGER PRIMARY KEY AUTOINCREMENT'],
                ['StudentID'     , 'TEXT UNIQUE DEFAULT NULL'],
                ['Email'         , 'TEXT'],
                ['Phone'         , 'TEXT'],
                ['Birthday'      , 'TEXT'],
                ['FormsDate'     , 'TEXT'],
                ['CampusLinkDate', 'TEXT'], ['MembStat', 'TEXT'], ['StudStat', 'TEXT'],
                ['Street'        , 'TEXT'], ['City'    , 'TEXT'], ['State'   , 'TEXT'], ['Zip'     , 'TEXT'],
                ['EmName'        , 'TEXT'], ['EmRel'   , 'TEXT'], ['EmPhoneH', 'TEXT'], ['EmPhoneW', 'TEXT'], ['EmPhoneC', 'TEXT'] ]
            for cert in Util.certifications:
                self.tableMemberDef += mkCertFields(cert)
            self.tableMemberDef += [
                ['RoommateName', 'TEXT'], ['RoommatePhone', 'TEXT'],
                ['InsurName'   , 'TEXT'], ['InsurPol'     , 'TEXT'], ['InsurGrp', 'TEXT'],
                ['Med'         , 'TEXT'], ['Note'         , 'TEXT'],
                ['LastUpdated' , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                 ', CONSTRAINT unq UNIQUE (FirstName, LastName, Birthday)']

            self.tableGearDef = ['Gear',
                ['Name'            , 'TEXT COLLATE NOCASE'],
                ['ID'              , 'TEXT COLLATE NOCASE'],
                ['Quantity'        , 'INTEGER DEFAULT 1'  ],
                ['Price'           , 'REAL DEFAULT 0.00'  ],
                ['Category'        , 'INTEGER DEFAULT 0'  ],
                ['Weight'          , 'REAL DEFAULT 0.0'   ],
                ['PurchaseDate'    , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['ExpirationDate'  , "TEXT DEFAULT ''"    ],
                ['Manufacturer'    , 'TEXT COLLATE NOCASE'],
                ['Unrentable'      , 'INTEGER DEFAULT 0'  ],
                ['UnrentableReason', "TEXT DEFAULT ''"    ]]
            for cert in Util.certifications:
                self.tableGearDef += [certDef(cert)]
            self.tableGearDef += [
                ['CareMaintenance', "TEXT DEFAULT ''"],
                ['Misc'           , "TEXT DEFAULT ''"],
                ['LastUpdated'    , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                 ', UNIQUE (ID)']

            self.tableRetiredGearDef = ['RetiredGear',
                ['Name'          , 'TEXT COLLATE NOCASE'],
                ['ID'            , 'TEXT COLLATE NOCASE'],
                ['Quantity'      , 'INTEGER DEFAULT 1'  ],
                ['Price'         , 'REAL DEFAULT 0.00'  ],
                ['Category'      , 'INTEGER DEFAULT 0'  ],
                ['Weight'        , 'REAL DEFAULT 0.0'   ],
                ['PurchaseDate'  , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['ExpirationDate', "TEXT DEFAULT ''"    ],
                ['Manufacturer'  , 'TEXT COLLATE NOCASE']]
            for cert in Util.certifications:
                self.tableRetiredGearDef += [certDef(cert)]
            self.tableRetiredGearDef += [
                ['CareMaintenance', "TEXT DEFAULT ''"],
                ['Misc'           , "TEXT DEFAULT ''"],
                ['RetiredDate'    , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"] ]

            self.tableTransactionDef = ['Transactions',
                ['TID'             , 'INTEGER PRIMARY KEY'],
                ['MID'             , 'INTEGER'            ],
                ['GID'             , 'TEXT'               ],
                ['CheckOutDateTime', 'TEXT'               ],
                ['DueDate'         , 'TEXT'               ],
                 ', FOREIGN KEY(MID) REFERENCES Member(ID), FOREIGN KEY(GID) REFERENCES Gear(ID)']

            self.tableArchiveTranDef = ['ArchiveTrans',
                ['MID_OUT'         , 'INTEGER'],
                ['MID_IN'          , 'INTEGER'],
                ['GID'             , 'TEXT'   ],
                ['CheckOutDateTime', 'TEXT'   ],
                ['DueDate'         , 'TEXT'   ],
                ['CheckInDateTime' , 'TEXT'   ],
                 ', FOREIGN KEY(MID_OUT) REFERENCES Member(ID), FOREIGN KEY(MID_IN) REFERENCES Member(ID), FOREIGN KEY(GID) REFERENCES Gear(ID)']

            self.tableFinancialTranDef = ['FinancialTrans',
                ['MID'    , 'INTEGER'],
                ['Date'   , 'TEXT'   ],
                ['Type'   , 'TEXT'   ],
                ['Amount' , 'REAL'   ],
                ['Comment', 'TEXT'   ],
                 ', FOREIGN KEY(MID) REFERENCES Member(ID)']

            self.tableGearMaintenanceDef = ['Maintenance',
                ['GID' , 'TEXT'],
                ['Date', 'TEXT'],
                ['Text', 'TEXT'],
                 ',FOREIGN KEY(GID) REFERENCES Gear(ID)']

            self.tableSettingsDef = ['Admin',
                ['DayOfMeetings', 'INTEGER DEFAULT 1'],
                ['MeetingFreq'  , 'INTEGER DEFAULT 1'],
                ['RentalFee'    , 'REAL DEFAULT 0.0' ],
                ['SemFallStart' , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['SemFallEnd'   , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['SemSprStart'  , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['SemSprEnd'    , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['SemSumStart'  , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"],
                ['SemSumEnd'    , "TEXT DEFAULT (date(CURRENT_TIMESTAMP, 'localtime'))"]]

            self.tableDefs = {self.tableMemberDef[0]: self.tableMemberDef,
                              self.tableGearDef[0]: self.tableGearDef,
                              self.tableTransactionDef[0]: self.tableTransactionDef,
                              self.tableArchiveTranDef[0]: self.tableArchiveTranDef,
                              self.tableSettingsDef[0]: self.tableSettingsDef,
                              self.tableFinancialTranDef[0]: self.tableFinancialTranDef,
                              self.tableGearMaintenanceDef[0]: self.tableGearMaintenanceDef,
                              self.tableRetiredGearDef[0]: self.tableRetiredGearDef}

            # self.execQuery("DROP TABLE IF EXISTS Member"        , 'database.py -> initDB, Member'        ); print 'Member dropped'
            # self.execQuery("DROP TABLE IF EXISTS Gear"          , 'database.py -> initDB, Gear'          ); print 'Gear dropped'
            # self.execQuery("DROP TABLE IF EXISTS Transactions"  , 'database.py -> initDB, Transactions'  ); print 'Transactions dropped'
            # self.execQuery("DROP TABLE IF EXISTS ArchiveTrans"  , 'database.py -> initDB, ArchiveTrans'  ); print 'ArchiveTrans dropped'
            # self.execQuery("DROP TABLE IF EXISTS Admin"         , 'database.py -> initDB, Admin'         ); print 'Admin dropped'
            # self.execQuery("DROP TABLE IF EXISTS FinancialTrans", 'database.py -> initDB, FinancialTrans'); print 'FinancialTrans dropped'
            # self.execQuery("DROP TABLE IF EXISTS Maintenance"   , 'database.py -> initDB, Maintenance'   ); print 'Maintenance dropped'

            # self.execQuery("ALTER TABLE Member ADD StudStat INTEGER DEFAULT 0", 'database.py -> initDB, ALTER TABLE')

            # Set settings
            self.execQuery('PRAGMA synchronous  = OFF'      , 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA journal_mode = MEMORY'   , 'database.py -> Database -> __init__')
#            self.execQuery('PRAGMA locking_mode = EXCLUSIVE', 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA auto_vacuum  = FULL'     , 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA page_size    = 65536'    , 'database.py -> Database -> __init__')

            global tableMember         ; tableMember = Table(self.tableMemberDef)
            global tableGear           ; tableGear = Table(self.tableGearDef)
            global tableTransaction    ; tableTransaction = Table(self.tableTransactionDef)
            global tableArchiveTrans   ; tableArchiveTrans = Table(self.tableArchiveTranDef)
            global tableAdmin          ; tableAdmin = Table(self.tableSettingsDef)
            global tableFinancialTran  ; tableFinancialTran = Table(self.tableFinancialTranDef)
            global tableGearMaintenance; tableGearMaintenance = Table(self.tableGearMaintenanceDef)
            global tableRetiredGear    ; tableRetiredGear = Table(self.tableRetiredGearDef)

            # Set settings
            self.execQuery('PRAGMA synchronous  = OFF'      , 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA journal_mode = MEMORY'   , 'database.py -> Database -> __init__')
#            self.execQuery('PRAGMA locking_mode = EXCLUSIVE', 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA auto_vacuum  = FULL'     , 'database.py -> Database -> __init__')
            self.execQuery('PRAGMA page_size    = 65536'    , 'database.py -> Database -> __init__')

            # Check to see if the admin table is empty, id so, initialize it with the default values
            query = self.getQuery('*', 'Admin')
            if self.qLen(query) <= 0:
                self.execQuery("INSERT INTO Admin (DayOfMeetings) VALUES (1)", 'database.py -> initDB, Admin')
            elif self.qLen(query) > 1:
                print 'ERROR: Admin table has more than one entry'
                raise ValueError

    def addItem(self, table_name, attrListIn):

        attr_list = []
        val_list = []
        table_def = self.tableDefs[table_name]

        for entry in table_def:
            if isinstance(entry, list) and len(entry) == 2:
                attr = entry[0]
                if attr in attrListIn.keys():
                    val = attrListIn[attr]

                    # Make sure that if StudentID is an empty string, it gets set to NULL
                    if attr == 'StudentID' and (not val or val.upper() == 'NULL'):
                        val = 'NULL'

                    if 'TEXT' in entry[1].upper() and not (attr == 'StudentID' and val == 'NULL'):
                        val = '"{0}"'.format(val.replace('"', '""'))

                    elif 'INTEGER' or 'REAL' in entry[1].upper():
                        # Convert from bool to int
                        if isinstance(val, bool):
                            val = int(val)
                        val = str(val)

                    attr_list.append(attr)
                    val_list.append(val)

        cmd = 'INSERT INTO {}({}) VALUES({})'.format(table_name, ', '.join(attr_list), ', '.join(val_list))

        self.execQuery(cmd, 'database.py -> addItem')

        return True

    def updateItem(self, table_name, attr_list, item):

        cmd = []
        tableDef = self.tableDefs[table_name]

        # Search for the type of the current attribute
        for entry in tableDef:
            if isinstance(entry, list) and len(entry) == 2:
                attr = entry[0]
                if attr in attr_list.keys():
                    val = attr_list[attr]

                    # Make sure that if StudentID is an empty string, it gets set to NULL
                    if attr == 'StudentID' and (val is None or val == '' or val.upper() == 'NULL'):
                        val = 'NULL'

                    # Add quotes around strings
                    if 'TEXT' in entry[1].upper() and not (attr == 'StudentID' and val == 'NULL'):
                        val = '"{0}"'.format(val.replace('"', '""'))

                    # Convert from bool to int
                    if isinstance(val, bool):
                        val = int(val)

                    cmd.append('{}={}'.format(attr, val))

        item_id = item.ID
        if table_name == 'Gear':
            item_id = '"{}"'.format(item_id)

        cmd = 'UPDATE {} SET {} WHERE ID={}'.format(table_name, ', '.join(cmd), item_id)

        self.execQuery(cmd, 'database.py -> updateItem')

        return True

    def delItem(self, table_name, item):

        id = item.ID
        if table_name == 'Gear':
            id = '"{}"'.format(id)

            self.archiveGear(item)

        cmd = 'DELETE FROM {} WHERE ID={}'.format(table_name, id)

        self.execQuery(cmd, 'database.py -> delItem')

    def wTransaction(self, member, gear, due_date):

        today_now = QtCore.QDateTime.currentDateTime().toString(Util.DB_DateTime)

        cmd = 'INSERT INTO Transactions(MID, GID, CheckOutDateTime, DueDate) VALUES({}, "{}", "{}", "{}")'.format(
            member.ID, gear.ID, today_now, due_date.toString(Util.DB_Date))

        self.execQuery(cmd, 'database.py -> wTransaction')

    def returnItem(self, trans, nTrans=1, retmemb=None):

        for n in xrange(nTrans):

            self.archiveTransaction(trans[n], retmemb)

            cmd = 'DELETE FROM {} WHERE TID={}'.format(tableTransaction.name, trans[n].TID)

            self.execQuery(cmd, 'database.py -> delTrans')

    def archiveTransaction(self, trans, retmemb=None):

        # Only archive if the item has been checked out for more than 1 hour
        today_now = QtCore.QDateTime.currentDateTime()
        checkOutDateTime = trans.CheckOutDateTime

        if checkOutDateTime.secsTo(today_now) >= 3600:

            today_now = today_now.toString(Util.DB_DateTime)

            if retmemb:
                return_id = retmemb.ID
            else:
                return_id = trans.MID

            cmd = ('INSERT INTO ArchiveTrans(MID_OUT, MID_IN, GID, CheckOutDateTime, DueDate, CheckInDateTime) '
                   'SELECT T.MID, {0} as MID_IN, T.GID, T.CheckOutDateTime, T.DueDate, "{1}" as CheckInDateTime '
                   'FROM Transactions T, Member M WHERE T.TID={2} AND T.MID=M.ID').format(return_id, today_now, trans.TID)

            self.execQuery(cmd, 'database.py -> archiveTransaction')

    def archiveGear(self, gear):

        # Only archive if the item has inventoried for more than a day
        today_now = QtCore.QDate.currentDate()
        check_out_date = gear.PurchaseDate

        if check_out_date.daysTo(today_now) >= 1:

            # today_now = today_now.toString(Util.DB_DateTime)

            certs = ''
            if Util.certifications:
                certs = ['{}Cert'.format(c) for c in Util.certifications]
                certs = ', ' + ', '.join(certs)

            cmd = ('INSERT INTO RetiredGear(Name, ID, Quantity, Price, Category, Weight, PurchaseDate, '
                   'ExpirationDate, Manufacturer{Cert}, CareMaintenance, Misc) '
                   'SELECT Name, ID, Quantity, Price, Category, Weight, PurchaseDate, '
                   'ExpirationDate, Manufacturer{Cert}, CareMaintenance, Misc '
                   'FROM Gear WHERE ID="{GID}"').format(Cert=certs, GID=gear.ID)

            self.execQuery(cmd, 'database.py -> archiveGear')

    def fillMember(self, name, memBDay, updCurFields=True):

        if updCurFields:
            self.parent.currentMemberNameID = name

        if name:

            first_name, last_name = self.split_name(name)

            if first_name and last_name:

                bDayList = self.getBDayList(first_name, last_name)

                oldBDayList = memBDay.get_menu()

                for n in xrange(len(oldBDayList)):
                    oldBDayList[n] = Util.convert_date('Disp2Qt', oldBDayList[n])

                if bDayList != oldBDayList:

                    if len(bDayList) > 1:
                        bDayList = [Util.selectOne] + bDayList

                    memBDay.setMenu(bDayList)

                return

        memBDay.clear()
        self.parent.currentBDayBoxIndex = 0

    def fill_gear(self, gearSearch, gearBox):

        gNameID = gearSearch.text()

        self.parent.currentGearLineEdit = gNameID

        if gNameID:

            query = None

            gear_name = [s for s in gearNameIDList[0] if gNameID.upper() == s.upper()]
            gear_id = [s for s in gearNameIDList[1] if gNameID.upper() == s.upper()]

            # Is gNameID in the Name column?
            if gear_name:

                query = self.getQuery('ID', 'Gear', 'Name', gNameID)

            # Is gNameID in the ID column?
            elif gear_id:

                query = self.getQuery('Name', 'Gear', 'ID', gNameID)

            if query and self.qLen(query) >= 1:

                item_list = self.query2list(query)

                old_item_list = gearBox.get_menu()

                if item_list != old_item_list:

                    if len(item_list) > 1:
                        item_list = [Util.selectOne] + item_list

                    gearBox.setMenu(item_list)

                return

        gearBox.clear()
        self.parent.currentGearComboBox = 0

    def isStudentIDAvailable(self, name, birthday, student_id):

        student_id = student_id.strip()

        # No ID is also unique
        if not student_id:
            return True

        query = self.getQuery('FirstName, LastName, Birthday', 'Member', 'StudentID', student_id)

        if self.qLen(query) == -1:
            return True
        elif self.qLen(query) == 1:
            first_name, last_name = self.split_name(name)
            query.first()
            query_bDay = Util.convert_date('DB2Qt', query.record().value(2))
            if first_name.upper() == query.record().value(0).upper() and last_name.upper() == query.record().value(1).upper() and birthday == query_bDay:
                return True

        return False

    def isGearIDUnique(self, gear_id):

        gear_id = gear_id.strip()

        query = self.getQuery('ID', 'Gear', 'ID', gear_id)

        return self.qLen(query) == -1

    def setDayOfMeetings(self, dayOfMeetings):

        self.setFieldTo('Admin', 'DayOfMeetings', dayOfMeetings)

    def setDueDate(self, due_date):

        self.setFieldTo('Admin', 'DueDate', Util.convert_date('Qt2DB', due_date))

    def get_day_of_meetings(self):

        cmd = 'SELECT DayOfMeetings FROM Admin'
        query = self.execQuery(cmd, 'database.py -> getDayOfMeetings')

        # A day was found in the table
        if self.qLen(query) == 1:

            query.first()

            return query.record().value(0)

    def getDefaultDueDate(self):

        day_of_meeting = self.get_day_of_meetings()

        days_until_due = day_of_meeting - QtCore.QDate.currentDate().dayOfWeek()

        if days_until_due <= 0:
            days_until_due += 7

        next_meeting_day = QtCore.QDate.currentDate().addDays(days_until_due)

        # If the next meeting day falls on a break, move it until after the break
        semester_dates = self.getSemesterDates()

        # While the next meeting is occurring between semesters,
        # push the date back 7 days
        if semester_dates['SemFallStart'].addDays(7) < semester_dates['SemFallEnd'] and \
                semester_dates['SemSprStart'].addDays(7) < semester_dates['SemSprEnd'] and \
                semester_dates['SemSumStart'].addDays(7) < semester_dates['SemSumEnd']:

            while next_meeting_day < semester_dates['SemFallStart'] or \
                    semester_dates['SemFallEnd'] < next_meeting_day < semester_dates['SemSprStart'] or \
                    semester_dates['SemSprEnd'] < next_meeting_day < semester_dates['SemSumStart']:
                next_meeting_day = next_meeting_day.addDays(7)

        every_n_weeks = self.get_meeting_frequency()

        next_meeting_day = next_meeting_day.addDays(7 * (every_n_weeks - 1))

        return next_meeting_day

    def get_meeting_frequency(self):

        return self.getAttrFromTable('Admin', 'MeetingFreq')

    def setMeetingFreq(self, meeting_frequency):

        self.setFieldTo('Admin', 'MeetingFreq', meeting_frequency)

    def getGearIDTransTable(self, row, member):

        cmd = 'SELECT T.GID FROM {} M, {} G, {} T WHERE M.ID={} AND M.ID=T.MID AND T.GID=G.ID ORDER BY T.DueDate ASC'.format(tableMember.name, tableGear.name, tableTransaction.name, member.ID)
        query = self.execQuery(cmd, 'database.py -> getGearTransTable')

        query.seek(row)

        return query.value(0)
        # return query.value(query.record().indexOf('T.GID'))

    def getAttrFromTable(self, table, attr, row=0, order=None):

        cmd = 'SELECT {} FROM {}'.format(attr, table)
        if order:
            cmd += " " + order

        query = self.execQuery(cmd, 'database.py -> getAttrFromTable')

        query.seek(row)

        return query.value(0)

    def getMID_GID_row_TransTable(self, row):

        cmd = 'SELECT GID, MID, TID FROM {}'.format(tableTransaction.name)
        query = self.execQuery(cmd, 'database.py -> getMID_GIDTransTable')

        query.seek(row)
        MID = query.value(query.record().indexOf('MID'))
        GID = query.value(query.record().indexOf('GID'))

        cmd = 'SELECT GID FROM {} WHERE MID={} ORDER BY TID ASC'.format(tableTransaction.name, MID)
        query = self.execQuery(cmd, 'database.py -> getMID_GID_row_TID_TransTable')
        while query.next():
            if query.value(0) == GID:
                row = query.at()
                break

        return MID, GID, row

    def getUniqueID(self, table, match):

        match = match.replace('*', '%')

        cmd = 'SELECT ID FROM {} WHERE ID LIKE "{}" ORDER BY ID ASC'.format(table, match)

        query = self.execQuery(cmd, 'database.py -> getUniqueID')

        if self.qLen(query) < 0:
            return match.replace('%', '1')
        else:
            dbID = []

            prefix, suffix = match.split('%')

            query.seek(-1)
            while query.next():
                ID = query.value(0)

                lenID = len(ID)

                # Generate a list of the number portion of the gear IDs that match the match_Orig string
                ID = ID.rstrip(suffix).lstrip(prefix)
                if Util.is_number(ID):
                    dbID.append(int(ID))

            # Iterate through the list of numbers.  If a gap is found in the number list, use that
            # missing number for the gearID.  If no gap is found, use the max+1 number for the gearID.
            newIDnum = None
            for n in xrange(1, len(dbID)):
                if dbID[n] > dbID[n - 1] + 1:
                    newIDnum = dbID[n - 1] + 1
                    break

            if not newIDnum:
                newIDnum = dbID[-1] + 1

            nZ = lenID - len(prefix) - len(suffix)
            return '{}{}{}'.format(prefix, str(newIDnum).zfill(nZ), suffix)

    def getBDayList(self, fName, lName):

        if fName and lName:

            query = self.getQuery('Birthday', 'Member', ['FirstName', 'LastName'], [fName, lName])

            if query and self.qLen(query) >= 1:

                bDayList = []

                query.seek(-1)
                while query.next():

                    bDayList.append(Util.convert_date('DB2Qt', query.value(0)))

                if bDayList:

                    return bDayList

    def set_rental_price(self, price):

        self.setFieldTo('Admin', 'RentalFee', price)

    def getRentalPrice(self):

        query = self.getQuery('RentalFee', 'Admin')

        query.first()

        return query.value(0)

    def submitPayment(self, payTypeBox, amountBox, payDateEdit, noteEdit):

        attr_list = {'MID': self.parent.currentMember.ID,
                     'Date': Util.convert_date('Qt2DB', payDateEdit.date()),
                     'Type': payTypeBox.currentText(),
                     'Amount': amountBox.value(),
                     'Comment': noteEdit.document().toPlainText().strip()}

        if attr_list['Type'] == 'Refund':
            attr_list['Amount'] = -attr_list['Amount']

        self.addItem('FinancialTrans', attr_list)

    def autoCompleter(self, findAttr, table, searchAtt=None, search=None, sort=None):

        item_list = {}
        attrs = findAttr.split(',')
        for n in xrange(len(attrs)):
            attrs[n] = attrs[n].strip()

            query = self.getQuery(attrs[n], table, searchAtt, search, sort)
            item_list[attrs[n]] = self.query2list(query, sort=False, ignore_duplicates=False)

        # Combine the names
        compList = []
        if 'FirstName' in findAttr and 'LastName' in findAttr and table == 'Member':
            first_name = item_list[attrs[0]]
            last_name = item_list[attrs[1]]
            for n in xrange(len(first_name)):
                compList.append('{0} {1}'.format(first_name[n], last_name[n]))
            for n in xrange(len(last_name)):
                compList.append('{0}, {1}'.format(last_name[n], first_name[n]))

            global fullFirstLast
            fullFirstLast = [compList, first_name + first_name, last_name + last_name]

        elif 'Name' in findAttr and 'ID' in findAttr and table == 'Gear':
            gear_name = Util.remove_duplicates(item_list[attrs[0]])
            gear_id = item_list[attrs[1]]

            compList = gear_name + gear_id

            global gearNameIDList
            gearNameIDList = [gear_name, gear_id]

        elif isinstance(findAttr, (basestring, unicode)):
            query = self.getQuery(findAttr, table, searchAtt, search, sort)
            compList = self.query2list(query)

        # Sort the list case insensitively
        compList = Util.sort_list(compList)

        completer = QtGui.QCompleter(compList)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)

        return completer

    def query2list(self, query, sort=True, ignore_duplicates=True):

        item_list = []
        for n in xrange(query.record().count()):
            query.seek(-1)
            while query.next():
                val = query.value(n)
                if val:
                    if isinstance(val, unicode):
                        val = str(val).strip()
                    item_list.append(val)

        # Sort the list case insensitively
        if len(item_list) > 1:

            # Remove duplicates
            if ignore_duplicates:

                item_list = list(set(item_list))

            if sort:
                if isinstance(item_list[0], (basestring, unicode)):
                    item_list = sorted(item_list, key=lambda s: s.lower())
                else:
                    print 'No sort available for this item type: database.py -> query2list'

        return item_list

    def getShitListQuery(self):

        cmd = ('SELECT M.FirstName, M.LastName, G.Name, T.GID, '
               'strftime("%m/%d/%Y %H:%M:%S", T.CheckOutDateTime), strftime("%m/%d/%Y", T.DueDate) '
               'FROM Transactions T JOIN Member M ON T.MID=M.ID JOIN Gear G ON T.GID=G.ID '
               'WHERE (JULIANDAY("NOW")-JULIANDAY(T.DueDate)) > 0 ORDER BY T.CheckOutDateTime DESC')

        query = self.execQuery(cmd, 'database.py -> getShitListQuery')

        return query

    def exportDatabase(self, file_name):

        wb = Workbook(guess_types=True)
        for sheet in wb:
            wb.remove_sheet(sheet)

        wb = self._make_sheet(tableDef=self.tableMemberDef         , title='Members'               , order='LastName'       , wb=wb)
        wb = self._make_sheet(tableDef=self.tableGearDef           , title='Gear Inventory'        , order='ID'             , wb=wb)
        wb = self._make_sheet(tableDef=self.tableTransactionDef    , title='Active Transactions'   , order='DueDate'        , wb=wb)
        wb = self._make_sheet(tableDef=self.tableFinancialTranDef  , title='Financial Transactions', order='Date'           , wb=wb)
        wb = self._make_sheet(tableDef=self.tableArchiveTranDef    , title='Archived Transactions' , order='CheckInDateTime', wb=wb)
        wb = self._make_sheet(tableDef=self.tableGearMaintenanceDef, title='Gear Maintenance'      , order='Date'           , wb=wb)
        wb = self._make_sheet(tableDef=self.tableSettingsDef       , title='Admin Settings'        , order='DayOfMeetings'  , wb=wb)
        wb = self._make_sheet(tableDef=self.tableRetiredGearDef    , title='Retired Gear'          , order='RetiredDate'    , wb=wb)

        wb.save(file_name)

    def exportShitList(self, file_name):

        wb = Workbook(guess_types=True)
        for sheet in wb:
            wb.remove_sheet(sheet)

        query = self.getShitListQuery()

        header = ['FirstName', 'LastName', 'Name', 'GID', 'strftime(%m/%d/%Y %H:%M:%S, T.CheckOutDateTime)', 'strftime(%m/%d/%Y, T.DueDate)']

        wb = self._make_sheet(header=header, query=query, title='Shit List', wb=wb)

        wb.save(file_name)

    def importDatabase(self, file_name):

        from openpyxl.cell import column_index_from_string

        wb = load_workbook(filename=file_name, use_iterators=True)
        ws = wb.get_sheet_by_name(name='Members')

        # Read and confirm the header
        for row in ws.iter_rows():
            membAttr = {}
            for cell in row:
                if cell._value:
                    col = column_index_from_string(cell.column)
                    if cell.row > 1:
                        # Read the remaining members
                        membAttr[self.impExpMemberHeader[col - 1]] = cell.value

                    elif cell.row == 1 and cell.value != self.impExpMemberHeader[col - 1]:
                        self.statBar.showMessage("Error in header in '{}'".format(file_name), Util.messageDelay)
                        return False

            if membAttr:
                member = self.getMember(membAttr['FirstName'], membAttr['LastName'], membAttr['Birthday'])
                if member:
                    self.updateItem('Member', membAttr, member)
                else:
                    self.addItem('Member', membAttr)

        return True

    def _make_sheet(self, **kwarg):

        for arg in kwarg.keys():

            if arg == 'tableDef':
                header = [v[0] for v in kwarg[arg] if isinstance(v, list) and len(v) == 2]
                table = kwarg[arg][0]
            elif arg == 'title':
                title = kwarg[arg]
            elif arg == 'order':
                order = kwarg[arg]
            elif arg == 'wb':
                wb = kwarg[arg]
            elif arg == 'query':
                query = kwarg[arg]
            elif arg == 'header':
                header = kwarg[arg]                

        ws = wb.create_sheet()
        ws.title = title

        # Write the header
        nameShift = 0
        for n, head in enumerate(header):
            c = ws.cell(row=1, column=n + 1 + nameShift)
            if head == 'MID' or head == 'MID_IN' or head == 'MID_OUT':
                c.value = 'LastName'
                if head == 'MID_IN' : c.value = 'Returned ' + c.value
                if head == 'MID_OUT': c.value = 'Checked Out  ' + c.value
                nameShift += 1
                c = ws.cell(row=1, column=n + 1 + nameShift)
                c.value = 'FirstName'
                if head == 'MID_IN' : c.value = 'Returned ' + c.value
                if head == 'MID_OUT': c.value = 'Checked Out  ' + c.value
            elif head == 'GID':
                c.value = 'ID'
                nameShift += 1
                c = ws.cell(row=1, column=n + 1 + nameShift)
                c.value = 'GearName'
            else:
                c.value = head.split('.')[-1].strip(')')

        if 'query' not in kwarg.keys():
            query = self.getQuery('*', table, sort='ORDER BY ' + order + ' ASC')

        # Dump the database
        row = 1
        while query.next():
            nameShift = 0
            for n, head in enumerate(header):

                dumpVal = query.record().value(query.record().indexOf(head))

                c = ws.cell(row=row + 1, column=n + 1 + nameShift)

                if head == 'MID' or head == 'MID_IN' or head == 'MID_OUT':
                    mNameQuery = self.getQuery("LastName, FirstName", 'Member', searchAtt='ID', search=str(dumpVal))
                    mNameQuery.first()
                    c.value = mNameQuery.value(0)
                    nameShift += 1
                    c = ws.cell(row=row + 1, column=n + 1 + nameShift)
                    c.value = mNameQuery.value(1)
                elif head == 'GID':
                    c.value = dumpVal
                    mNameQuery = self.getQuery("Name", 'Gear', searchAtt='ID', search=str(dumpVal))
                    mNameQuery.first()
                    nameShift += 1
                    c = ws.cell(row=row + 1, column=n + 1 + nameShift)
                    c.value = mNameQuery.value(0)
                else:
                    c.value = dumpVal

                if isinstance(c.value, (basestring, unicode)) and c.value.count('-') == 2:
                    y, m, d = dumpVal.split('-')
                    if len(y) == 4 and len(m) == 2 and len(d) == 2:
                        c.value = datetime.datetime(int(y), int(m), int(d))
                        c.number_format = Util.dateDispFormat

            row += 1

        return wb

    def close(self):

        # Reset the connector and remove it (magic)
        self.SQLDB = QtSql.QSqlDatabase()
        self.execQuery('VACUUM', 'Database.py -> Database -> close')
        self.SQLDB.close()
        self.SQLDB.removeDatabase(self.nameDB)
        del self.SQLDB

        # Backup the database file
        backupDir = 'Backup_Databases'
        if not os.path.isdir(backupDir):
            os.mkdir(backupDir)
        shutil.copyfile(self.nameDB, '{}/{}_{}'.format(backupDir,
                                                       QtCore.QDate.currentDate().toString('yyMMdd'),
                                                       self.nameDB))
